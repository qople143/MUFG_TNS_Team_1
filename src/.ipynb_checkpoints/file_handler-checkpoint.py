import pandas as pd
from openpyxl import load_workbook


def validate_excel_file(file_path):
    try:
        if not file_path.endswith('.xlsx'):
            raise ValueError("File must be .xlsx format")
        
        wb = load_workbook(file_path, data_only=True)
        if len(wb.sheetnames) == 0:
            raise ValueError("Excel file has no sheets")
        
        return True
    except Exception as e:
        raise ValueError(f"File validation failed: {str(e)}")


def load_excel_file(file_path, sheet_name=0):
    try:
        if isinstance(sheet_name, int):
            excel_file = pd.ExcelFile(file_path)
            sheet_name = excel_file.sheet_names[sheet_name]
        
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {str(e)}")


def get_sheet_names(file_path):
    try:
        excel_file = pd.ExcelFile(file_path)
        return excel_file.sheet_names
    except Exception as e:
        raise ValueError(f"Failed to read sheets: {str(e)}")


def save_excel_file(df, output_path, sheet_name='Transformed'):
    try:
        df.to_excel(output_path, sheet_name=sheet_name, index=False)
        return True
    except Exception as e:
        raise ValueError(f"Failed to save Excel file: {str(e)}")


def preview_data(df, rows=5):
    return df.head(rows)